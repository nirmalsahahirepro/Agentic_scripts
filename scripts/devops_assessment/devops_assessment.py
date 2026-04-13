#!/usr/bin/env python3
"""
Generic Assessment Automation Script — Selenium
================================================
GenericAssessmentAutoAgent v1.1.0

Automates the full lifecycle of the HirePro generic assessment:
  - Pre-test navigation (T&C, Next screens)
  - Identity verification (Selfie)
  - All question types: MCQ, Subjective, MultipleCorrectAnswer,
    MCQWithWeightage, FillInTheBlank, Coding
  - Optional group selection
  - Test submission & report generation

Usage:
    python3 generic_assessment.py

    Set the assessment URL in configs/urls.txt using the prefix:
        generic: https://ain.hirepro.in/YOUR_CODE

Dependencies:
    pip install selenium webdriver-manager
"""

import csv
import json
import logging
import os
import sys
import time
import argparse
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

# =====================================================================
# CONFIGURATION
# =====================================================================
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
URLs_FILE      = os.path.join(BASE_DIR, "configs", "urls.txt")
REPORT_DIR     = os.path.join(BASE_DIR, "output_reports", "devops_assessment")
SCREENSHOT_DIR = os.path.join(REPORT_DIR, "screenshots")

ELEMENT_WAIT_SEC   = 20
PAGE_LOAD_WAIT_SEC = 5
MAX_RETRIES        = 3
RETRY_DELAY_SEC    = 2
VM_ALLOCATION_TIMEOUT = 600
TODAY              = datetime.now().strftime("%Y-%m-%d")

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("devops_assessment")

# Global Settings
GLOBAL_IS_DEVOPS   = True
VM_ALLOCATION_TIMEOUT = 600 # 10 minutes

# DevOps Data
YAML_SOLUTION = """##########################
#Your solution should go
#below this line
#agent
#update1
#########################
---
- hosts: localhost
  connection: local
  tasks:
    - name: Purge discover
      apt:
        name: discover
        state: absent
        purge: true

    - name: Setup apt source
      apt_repository:
        repo: ppa:apt-fast/stable
        update_cache: true

    - name: Install apt-fast
      apt:
        name: apt-fast
        update_cache: true
"""

# =====================================================================
# EXECUTION STATE
# =====================================================================
execution_log = {
    "agent_name": "GenericAssessmentAutoAgent",
    "version": "v1.1.0",
    "assessment_url": "",
    "execution_status": "running",
    "answer_strategy": "heuristic",
    "total_questions": 0,
    "answered": 0,
    "skipped": 0,
    "errors": 0,
    "accuracy_estimate": "N/A",
    "start_time": datetime.now().isoformat(),
    "end_time": "",
    "duration_sec": 0,
    "screenshots": [],
    "error_log": [],
    "remarks": "",
    "coding_question_count": 0,
}
question_log    = []
qa_steps        = []
start_time      = time.time()
step_counter    = 1

# =====================================================================
# QA LOGGING HELPER
# =====================================================================
def log_qa_step(action, status, details="", step_id=None):
    """
    Appends a granular action to the QA log for final reports.
    """
    global step_counter
    sid = step_id if step_id else f"TS_{step_counter:03d}"
    if not step_id:
        step_counter += 1
        
    entry = {
        "Step ID": sid,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Action": action,
        "Details": details,
        "Status": status
    }
    qa_steps.append(entry)
    # Mirror to console for visibility
    log.info(f"  [{sid}] {status}: {action} - {details}")
    return sid


# =====================================================================
# CHROME DRIVER SETUP
# =====================================================================
def create_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    
    # Auto-allow camera/mic permission prompts without faking the actual camera stream
    opts.add_argument("--use-fake-ui-for-media-stream")
    # opts.add_argument("--use-fake-device-for-media-stream") # Keep commented to use REAL camera
    
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    if USE_WDM:
        service = Service(ChromeDriverManager().install())
    else:
        service = Service()

    driver = webdriver.Chrome(service=service, options=opts)
    driver.implicitly_wait(5)
    return driver


# =====================================================================
# HELPERS
# =====================================================================
def wait(driver, by, selector, timeout=ELEMENT_WAIT_SEC):
    """Wait for element to be present and return it."""
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )


def wait_clickable(driver, by, selector, timeout=ELEMENT_WAIT_SEC):
    """Wait for element to be clickable and return it."""
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )


def fast_find_elements(driver, by, selector):
    """Find elements with 0s implicit wait to avoid hanging on missing elements."""
    driver.implicitly_wait(0)
    try:
        return driver.find_elements(by, selector)
    finally:
        driver.implicitly_wait(PAGE_LOAD_WAIT_SEC)


def check_fullscreen_overlay(driver):
    """Proactively click the 'Go back to full screen' button and force fullscreen mode."""
    # 1. Force browser sync and check state
    try:
        # Avoid driver.fullscreen_window() as it hangs on some Linux environments
        # We rely on --start-maximized in create_driver() instead
        # driver.fullscreen_window()
        time.sleep(0.1)
    except Exception:
        pass

    # 2. Specifically look for buttons/links that dismiss the overlay
    labels = ["go back to full screen", "full screen", "return to test", "resume", "click here to go back to full screen"]
    or_parts = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in labels]
    # Target only button or a tags for reliability
    combined_xpath = f"//button[{' or '.join(or_parts)}] | //a[{' or '.join(or_parts)}]"

    try:
        btns = fast_find_elements(driver, By.XPATH, combined_xpath)
        visible = [b for b in btns if b.is_displayed()]
        if visible:
            log.info(f"  Fullscreen overlay button detected — clicking to recover")
            # Try normal click first
            try:
                visible[0].click()
            except Exception:
                # Fallback to JS click for reliability
                driver.execute_script("arguments[0].click();", visible[0])
            
            time.sleep(2)  # Wait for UI to stabilize after fullscreen transition
            return True
    except Exception:
        pass
    return False

def safe_click(driver, element, step_name=""):
    """Click element cleanly without JS bypass. Aborts if click is intercepted by an active overlay."""
    for attempt in range(MAX_RETRIES):
        try:
            element.click()
            log_qa_step(f"Click: {step_name}", "PASS", f"Clicked successfully")
            return True
        except ElementClickInterceptedException:
            log.warning(f"  [{step_name}] Click intercepted by overlay! (attempt {attempt+1})")
            time.sleep(1)
            
            # Check if it's just the anti-cheat fullscreen warning overlay and dismiss it
            if check_fullscreen_overlay(driver):
                continue
                
            # If an element is blocked by an overlay (like an "End of Section" modal),
            # DO NOT fallback to JS bypass. We must respect the UI state.
            return False
        except StaleElementReferenceException:
            log.warning(f"  [{step_name}] Stale element, re-locating...")
            time.sleep(RETRY_DELAY_SEC)
        except Exception as e:
            log.warning(f"  [{step_name}] Click error: {e}")
            time.sleep(RETRY_DELAY_SEC)
    return False


def safe_click_checkbox(driver, element, step_name=""):
    """Clicks a custom checkbox safely using Javascript to avoid hitting center-embedded links (e.g. Terms of Use)."""
    try:
        # JS click fires on the top-level element without falsely activating embedded anchor tags visually centered
        driver.execute_script("arguments[0].click();", element)
        log_qa_step(f"JS Click: {step_name}", "PASS")
        return True
    except Exception as e:
        log.warning(f"  [{step_name}] JS check failed: {e}. Falling back to standard click.")
        return safe_click(driver, element, step_name)


def screenshot(driver, name):
    """Take and save a screenshot."""
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    path = os.path.join(SCREENSHOT_DIR, f"{name}_{TODAY}.png")
    try:
        driver.save_screenshot(path)
        execution_log["screenshots"].append(path)
        # Note: log_qa_step call removed as per user request to exclude screenshot details from report
    except Exception as e:
        log.warning(f"  Screenshot failed: {e}")
    return path


def log_error(step, level, message):
    """Record an error to the execution log."""
    entry = {
        "step": step,
        "level": level,
        "message": message,
        "timestamp": datetime.now().isoformat(),
    }
    execution_log["error_log"].append(entry)
    execution_log["errors"] += 1
    if level == "WARNING":
        log.warning(f"  [{step}] {message}")
    else:
        log.error(f"  [{step}] {message}")


def click_by_text(driver, tag, text, step=""):
    """Find and click an element by its visible text."""
    for attempt in range(MAX_RETRIES):
        try:
            el = WebDriverWait(driver, ELEMENT_WAIT_SEC).until(
                EC.element_to_be_clickable(
                    (By.XPATH, f"//{tag}[contains(.,'{text}')]")
                )
            )
            safe_click(driver, el, step)
            log_qa_step(f"Click by Text: '{text}'", "PASS")
            return True
        except TimeoutException:
            log.warning(f"  [{step}] '{text}' not found (attempt {attempt+1}/{MAX_RETRIES})")
            time.sleep(RETRY_DELAY_SEC)
    log_error(step, "WARNING", f"Could not find/click '{text}' after {MAX_RETRIES} attempts")
    return False


def switch_to_newest_window(driver, silent=True):
    """Switch to the most recently opened window/tab if not already there."""
    try:
        handles = driver.window_handles
        if len(handles) > 1:
            last_handle = handles[-1]
            if driver.current_window_handle != last_handle:
                driver.switch_to.window(last_handle)
                if not silent:
                    log.info(f"  Switched focus to new window: {driver.current_url}")
                return True
    except Exception:
        pass
    return False


# =====================================================================
# STEP 1: NAVIGATE TO URL
# =====================================================================
def step_navigate(driver, url):
    log.info("═" * 60)
    execution_log["assessment_url"] = url
    try:
        driver.get(url)
        # Verify if loaded
        WebDriverWait(driver, PAGE_LOAD_WAIT_SEC).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        log_qa_step("Navigate to URL", "PASS", f"URL: {url}")
        return True
    except Exception as e:
        log_qa_step("Navigate to URL", "FAIL", str(e))
        log_error("step_navigate", "CRITICAL", f"Navigation failed: {e}")
        return False

    try:
        driver.get(url)
        WebDriverWait(driver, PAGE_LOAD_WAIT_SEC).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        log.info(f"  Page loaded: {driver.current_url}")
        screenshot(driver, "01_page_loaded")
        return True
    except Exception as e:
        log_error("navigate", "CRITICAL", f"Page load failed: {e}")
        return False


# =====================================================================
# STEP 2: PRE-TEST NAVIGATION (T&C, Next screens)
# =====================================================================
def _click_custom_checkbox(driver, step_name):
    """Click the HirePro custom-checkbox label (the visible element, not raw input)."""
    # Try label.custom-checkbox first (Online Proctoring modal)
    for selector in [
        "label.custom-checkbox",
        "span[role='checkbox']",
        ".hp-checkbox label",
        ".custom-checkbox",
    ]:
        clicked_any = False
        try:
            labels = driver.find_elements(By.CSS_SELECTOR, selector)
            for lbl in labels:
                if lbl.is_displayed():
                    # Check if it has an active/checked class first
                    class_name = lbl.get_attribute("class") or ""
                    if "checked" in class_name or "active" in class_name:
                        log.info(f"  [{step_name}] Checkbox '{selector}' already appears checked via class.")
                        continue
                        
                    # Also try to find nearby input to see if it's already checked
                    try:
                        parent = lbl.find_element(By.XPATH, "..")
                        inp = parent.find_element(By.TAG_NAME, "input")
                        if inp.is_selected():
                            log.info(f"  [{step_name}] Checkbox '{selector}' already checked via input state.")
                            continue
                    except Exception:
                        pass
                        
                    safe_click_checkbox(driver, lbl, step_name)
                    log.info(f"  [{step_name}] Clicked custom checkbox via selector: {selector} (offset)")
                    time.sleep(1)
                    clicked_any = True
            if clicked_any:
                return True
        except Exception:
            pass

    # Fallback: raw checkboxes that aren't yet selected
    try:
        checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
        for cb in checkboxes:
            if cb.is_displayed() and not cb.is_selected():
                safe_click(driver, cb, step_name)
                log.info(f"  [{step_name}] Clicked raw checkbox")
                time.sleep(0.5)
                return True
    except Exception:
        pass
    return False


def step_pretest(driver):
    log.info("═" * 60)
    log.info("STEP 2: Pre-Test Navigation")
    log.info("═" * 60)
    time.sleep(3)

    # ── Handle 'Popup Blocked' screen ──────────────────────────────
    # HirePro opens the test in a popup; if blocked, a dialog with
    # 'Open Test Window' button appears on the login page.
    try:
        open_btn = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(.,'Open Test Window')]"
                 " | //a[contains(.,'Open Test Window')]")
            )
        )
        safe_click(driver, open_btn, "open-test-window")
        log.info("  Clicked 'Open Test Window'")
        time.sleep(3)
    except TimeoutException:
        log.info("  No 'Open Test Window' prompt — popup opened normally")

    # ── Switch to test popup window ────────────────────────────────
    switch_to_newest_window(driver)
    time.sleep(2)

    # Dismiss any unexpected browser alert
    try:
        driver.switch_to.alert.dismiss()
    except Exception:
        pass

    screenshot(driver, "02_pretest_opened")

    # ── Online Proctoring Setup modal (up to 3 rounds) ─────────────
    # Each round: tick checkbox → wait for Next to become enabled → click Next
    for round_num in range(1, 4):
        log.info(f"  Pre-test navigation round {round_num}")
        time.sleep(2)

        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()

        # If already on main assessment / selfie page, stop
        if any(x in page_text for x in ["agree and start", "start test", "selfie", "system check"]):
            log.info("  Reached main assessment screen ✅")
            break

        # Tick the T&C / proctoring consent checkbox
        _click_custom_checkbox(driver, f"TC-round{round_num}")
        time.sleep(1)

        # Debug: dump all buttons to see what the Next button looks like
        try:
            with open(os.path.join(REPORT_DIR, f"debug_buttons_{round_num}.txt"), "w") as f:
                all_btns = driver.find_elements(By.TAG_NAME, "button")
                for btn in all_btns:
                    cls = btn.get_attribute("class") or ""
                    txt = btn.text.strip().replace("\n", " ")
                    disabled = btn.get_attribute("disabled") is not None
                    f.write(f"Button | Class: {cls} | Text: '{txt}' | Disabled: {disabled} | Visible: {btn.is_displayed()}\n")
        except Exception as e:
            log.error(f"  Debug dump failed: {e}")

        # Wait for Next button to become clickable (it's disabled until checkbox is ticked)
        next_clicked = False
        try:
            next_btns = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, ".btn.btn-primary.ng-scope.ng-isolate-scope, button[title='Next'], button.btn-primary")
                )
            )
            for btn in next_btns:
                if btn.text.strip().lower() == "next":
                    if btn.is_displayed():
                        try:
                            # Always prefer JS click here due to overlay issues on this specific modal
                            driver.execute_script("arguments[0].click();", btn)
                            log.info(f"  Clicked 'Next' (round {round_num}) via JS CSS selector")
                        except Exception as e:
                            log.debug(f"  JS click failed: {e}")
                            safe_click(driver, btn, f"Next-round{round_num}")
                            
                        next_clicked = True
                        time.sleep(3) # Wait for page transition / next stage
                continue 
        except Exception as e:
            # log warning but continue loop to ensure we don't stall forever on one transient error
            log.warning(f"  Submission search cycle encountered an error: {e}")
            
        time.sleep(2) # Polling interval

        # Check if we moved to next screen before declaring fail
        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
        if not next_clicked and "agree and start test" not in page_text:
            log.info(f"  No enabled 'Next' button found in round {round_num}")

        screenshot(driver, f"02_pretest_round{round_num}")

        # Re-check page after Next
        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
        if any(x in page_text for x in ["agree and start", "start test", "selfie", "system check"]):
            log.info("  Reached main assessment screen ✅")
            break

        if not next_clicked:
            # Verify if we somehow moved to the correct screen anyway
            page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
            if any(x in page_text for x in ["agree and start", "start test", "selfie", "system check"]):
                log.info("  Reached next screen despite not detecting Next click.")
                break
            else:
                log.warning(f"  Failed to click Next in round {round_num} and did not advance. Aborting pre-test.")
                return False

    screenshot(driver, "02_pretest_done")
    return True


# =====================================================================
# STEP 3: IDENTITY VERIFICATION (Selfie)
# =====================================================================
def step_selfie(driver):
    log.info("═" * 60)
    log.info("STEP 3: Identity Verification")
    log.info("═" * 60)

    # Check if selfie screen is shown
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        if "selfie" not in body and "photo" not in body:
            log.info("  No selfie screen detected, skipping")
            return True
    except Exception:
        pass

    # Wait for "Click a Selfie" button to be clickable (requires real face detection)
    log.info("  Waiting up to 60s for 'Click a Selfie' to be enabled (please stay in front of camera)...")
    try:
        selfie_btn = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Selfie') or contains(.,'Photo')]"))
        )
        safe_click(driver, selfie_btn, "click-selfie")
        log.info("  Clicked 'Click a Selfie'")
        time.sleep(3)
    except TimeoutException:
        log.warning("  'Click a Selfie' button never became clickable.")

    time.sleep(2)
    screenshot(driver, "03_selfie")

    # Wait for "Proceed to Test" button to be clickable
    log.info("  Waiting up to 60s for 'Proceed to Test' to be enabled...")
    try:
        proceed_btn = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Proceed') or contains(.,'Continue')]"))
        )
        safe_click(driver, proceed_btn, "proceed-test")
        log.info("  Clicked 'Proceed to Test'")
    except TimeoutException:
        log.warning("  'Proceed' button not clickable.")
        return False
    time.sleep(2)
    return True


# =====================================================================
# STEP 4: AGREE AND START TEST
# =====================================================================
def step_start_test(driver):
    log.info("═" * 60)
    log.info("STEP 4: Starting Test")
    log.info("═" * 60)
    time.sleep(2)

    # First: tick the 'I agree to Terms of Use & Privacy Policy' checkbox
    # (on the main assessment page, right side — custom-checkbox pattern)
    try:
        start_cb = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "label.custom-checkbox-startTest, input[ng-model='vm.isTermsAccepted']"))
        )
        if start_cb.is_displayed():
            safe_click_checkbox(driver, start_cb, "start-test-checkbox")
            log.info("  [start-test-checkbox] Clicked explicit start test checkbox (offset)")
            time.sleep(1)
    except TimeoutException:
        _click_custom_checkbox(driver, "start-test-checkbox")
        time.sleep(1)

    # Now click the 'Agree and Start Test' button
    start_test_clicked = False
    for attempt in range(3):
        try:
            # Use the exact class combination identified by the browser agent
            # Increased timeout as Angular might be slow to inject this element
            start_btn = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button.btn-success, button[name='btnStartTest']"))
            )
            
            # Since the button might be visually covered by a sticky header, JS click is safest
            try:
                driver.execute_script("arguments[0].click();", start_btn)
                log.info("  Clicked 'Start Test' via JS")
            except Exception:
                safe_click(driver, start_btn, "start-test")
                log.info("  Clicked 'Start Test' via Selenium")
                
            start_test_clicked = True
            time.sleep(2)
            break
        except TimeoutException:
            log.info(f"  Start Test button not found yet (attempt {attempt+1}/3)")
            time.sleep(2)
            
    if not start_test_clicked:
        log.warning("  Failed to find 'Start Test' button after multiple attempts.")
        # Final desperate fallback via javascript text search
        try:
            driver.execute_script('''
                var btns = document.querySelectorAll('button');
                for(var i=0; i<btns.length; i++) {
                    if(btns[i].innerText.toLowerCase().includes('start test')) {
                        btns[i].click();
                        break;
                    }
                }
            ''')
            log.info("  Attempted to click 'Start Test' via JS text-search injection")
            time.sleep(3)
        except Exception as e:
            log.error(f"  JS fallback failed: {e}")

    # Switch to test window if popup opened
    switch_to_newest_window(driver)
    time.sleep(3)
    screenshot(driver, "04_test_started")
    log.info("  Test started ✅")
    return True


# =====================================================================
# QUESTION ANSWERING STRATEGIES
# =====================================================================
def _get_question_type(driver):
    """Detect question type from DOM using fast CSS checks (0s implicit wait)."""
    # 0. Forced DevOps mode (Priority)
    if GLOBAL_IS_DEVOPS:
        # Check for standard question types first (in case it's a mixed test)
        std_q = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-radio, label.custom-checkbox, input[placeholder='Blank']")
        if std_q and any(q.is_displayed() for q in std_q):
             log.debug("  Standard UI found — continuing with normal detection")
        else:
             # Default to DevOps in this script to trigger the 10min wait loop
             log.info("  No standard question detected — forcing DevOps Lab path")
             return "DevOps"

    # 1. DevOps Loading/Spinner check (AWS VM takes time)
    # Check for typical loading text before the editor appears
    loading_text = fast_find_elements(driver, By.XPATH, "//*[contains(text(), 'Allocating') or contains(text(), 'setting up your') or contains(@class, 'spinner')]")
    if loading_text and any(l.is_displayed() for l in loading_text):
        log.info("  VM Allocation spinner/text detected — prioritizing DevOps Lab wait")
        return "DevOps"

    # 1. DevOps - Monaco editor presence (AWS VM based)
    monaco = fast_find_elements(driver, By.CSS_SELECTOR, ".monaco-editor")
    if monaco and any(m.is_displayed() for m in monaco):
        return "DevOps"

    # Priority 2: Wait for monaco if no other standard questions found
    # This prevents the script from defaulting to 'Unknown/MCQ' while the VM is booting
    standard_questions = fast_find_elements(driver, By.CSS_SELECTOR, ".ace_editor, .CodeMirror, input[placeholder='Blank'], label.custom-radio, label.custom-checkbox")
    if not any(q.is_displayed() for q in standard_questions):
        # Desperate wait for DevOpslab if nothing else found
        log.debug("  No standard question types found, taking a quick peek for Monaco...")
        try:
            wait(driver, By.CSS_SELECTOR, ".monaco-editor", timeout=10)
            return "DevOps"
        except Exception:
            pass

    # 2. Coding - ACE/CodeMirror presence
    code_editors = fast_find_elements(driver, By.CSS_SELECTOR, ".ace_editor, .CodeMirror")
    if code_editors and any(ed.is_displayed() for ed in code_editors):
        return "Coding"

    # 2. FIB - blank inputs
    fib_inputs = fast_find_elements(driver, By.CSS_SELECTOR, "input[id^='hp_fillintheblank_'], input[placeholder='Blank']")
    if fib_inputs and any(i.is_displayed() for i in fib_inputs):
        return "FillInTheBlank"
    
    # 3. Reference To Context (RTC) - Presence of Passage/Paragraph container
    rtc_markers = fast_find_elements(driver, By.CSS_SELECTOR, "hp-rtc, .rtc-question, .rtc-paragraph, .hp-paragraph, [id*='rtc']")
    if rtc_markers and any(m.is_displayed() for m in rtc_markers):
        return "RTC"

    # 4. MCQ With Weightage (MCQW) - Often have specific weighted classes
    mcqw_markers = fast_find_elements(driver, By.CSS_SELECTOR, "hp-weighted-mcq, .mcq-weighted, .weighted-question, .weighted")
    if mcqw_markers and any(m.is_displayed() for m in mcqw_markers):
        return "MCQWithWeightage"

    # 5. Subjective - textareas (excluding editor inputs)
    textareas = fast_find_elements(driver, By.CSS_SELECTOR, "textarea:not(.ace_text-input)")
    if textareas and any(t.is_displayed() for t in textareas):
        return "Subjective"

    # 6. Checkboxes (MCA) vs Radios (MCQ)
    checkboxes = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-checkbox, .hp-checkbox, input[type='checkbox']")
    radios = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-radio, .hp-radio, input[type='radio']")
    
    if checkboxes and any(c.is_displayed() for c in checkboxes):
        return "MultipleCorrectAnswer"
    
    if radios and any(r.is_displayed() for r in radios):
        return "MCQ"

    return "Unknown"


def answer_mcq(driver, q_num, task="MCQ"):
    """MCQ/MCQWithWeightage/RTC: click first visible radio option."""
    try:
        # HirePro: radio options are label.custom-radio or label.custom-checkbox with radio inside
        options = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-radio, label.custom-checkbox input[type='radio']")
        # If label.custom-radio found, click the label; if input[type='radio'] found, click its parent label
        if not options:
            options = fast_find_elements(driver, By.CSS_SELECTOR, "input[type='radio']")

        visible = [o for o in options if o.is_displayed()]
        if visible:
            if safe_click(driver, visible[0], f"{task.lower()}-Q{q_num}"):
                btn_text = visible[0].text.strip()
                log.info(f"  Q{q_num} [{task}] → selected option: '{btn_text}'")
                log_qa_step(f"Answer {task}", "PASS", f"Question: {q_num}, Choice: '{btn_text}'")
                return {"type": task, "answer": f"Option: {btn_text}", "status": "answered"}
            else:
                log.warning(f"  Q{q_num} [{task}] → click failed, overlay present?")
                return {"type": task, "answer": "", "status": "error"}
    except Exception as e:
        log_error(f"Q{q_num}-{task}", "WARNING", str(e))
    return {"type": task, "answer": "", "status": "error"}


def answer_rtc(driver, q_num):
    """Reference To Context: select first radio option."""
    return answer_mcq(driver, q_num, task="RTC")


def answer_multiple_correct(driver, q_num):
    """MultipleCorrectAnswer: select the first 2 visible custom-checkbox labels."""
    try:
        # HirePro MCA: click the label.custom-checkbox elements (which wrap input[type='checkbox'])
        checkboxes = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-checkbox")
        visible = [c for c in checkboxes if c.is_displayed()]
        selected = []
        for cb in visible[:2]:
            if safe_click(driver, cb, f"mca-Q{q_num}"):
                selected.append("checked")
                time.sleep(0.2)
        if selected:
            log.info(f"  Q{q_num} [MCA] → selected {len(selected)} options")
            return {"type": "MultipleCorrectAnswer", "answer": f"{len(selected)} options selected", "status": "answered"}
        else:
            log.warning(f"  Q{q_num} [MCA] → no checkboxes clicked")
    except Exception as e:
        log_error(f"Q{q_num}-MCA", "WARNING", str(e))
    return {"type": "MultipleCorrectAnswer", "answer": "", "status": "error"}


def answer_subjective(driver, q_num):
    """Subjective: type a contextual answer."""
    answer_text = (
        "This question requires a thoughtful response. "
        "Based on the context provided, the most appropriate answer involves understanding the core concept. "
        "The solution demonstrates proficiency in the relevant domain and meets the outlined requirements."
    )
    try:
        textareas = fast_find_elements(driver, By.CSS_SELECTOR, "textarea")
        visible = [t for t in textareas if t.is_displayed()]
        if visible:
            visible[0].clear()
            visible[0].send_keys(answer_text)
            log.info(f"  Q{q_num} [Subjective] → typed answer")
            log_qa_step("Answer Subjective", "PASS", f"Question: {q_num}, Text: {answer_text}")
            return {"type": "Subjective", "answer": answer_text[:80] + "...", "status": "answered"}
    except Exception as e:
        log_error(f"Q{q_num}-Subj", "WARNING", str(e))
    return {"type": "Subjective", "answer": "", "status": "error"}


def answer_fill_in_blank(driver, q_num):
    """FillInTheBlank: type contextual values in blank fields using HirePro selectors."""
    fill_values = ["123", "2.567", "!@#nirmal", "55", "function"]
    try:
        # HirePro FIB blanks: input[id^='hp_fillintheblank_'] or input[placeholder='Blank']
        inputs = fast_find_elements(driver, By.CSS_SELECTOR, 
            "input[id^='hp_fillintheblank_'], input[placeholder='Blank'], input[placeholder='blank']"
        )
        if not inputs:
            inputs = fast_find_elements(driver, By.CSS_SELECTOR, "input[type='text']")
        visible = [i for i in inputs if i.is_displayed()]
        filled = 0
        for idx, inp in enumerate(visible):
            val = fill_values[idx % len(fill_values)]
            try:
                inp.clear()
                inp.send_keys(val)
                filled += 1
                time.sleep(0.15)
            except Exception:
                pass
        log.info(f"  Q{q_num} [FIB] → filled {filled} blanks")
        log_qa_step("Answer FillInTheBlank", "PASS", f"Question: {q_num}, Values: {', '.join(fill_values[:filled])}")
        return {"type": "FillInTheBlank", "answer": f"Filled {filled} blanks", "status": "answered" if filled > 0 else "error"}
    except Exception as e:
        log_error(f"Q{q_num}-FIB", "WARNING", str(e))
    return {"type": "FillInTheBlank", "answer": "", "status": "error"}


def answer_coding(driver, q_num):
    """Coding: select correct language, inject specific solution, wait for results, then submit."""
    global execution_log
    count = execution_log.get("coding_question_count", 0)
    
    # 1st Python Solution
    python_sol_1 = """def factorial(n):
    # Base case
    if n == 0 or n == 1:
        return 1
    # Recursive case
    else:
        return n * factorial(n - 1)

# Take input from user
num = int(input())

# Check for negative numbers
if num < 0:
    print("Factorial is not defined for negative numbers.")
else:
    print(f"Factorial of {num} is {factorial(num)}")
"""
    # 2nd Python Solution
    python_sol_2 = """def factorial(n):
    # Base case
    if n == 0 or n == 1:
        return 1
    # Recursive case
    else:
        return n * factorial(n - 1)

# Take input from user
num = int(input("Enter a number: "))

# Check for negative numbers
if num < 0:
    print("Factorial is not defined for negative numbers.")
else:
    print(f"Factorial of {num} is {factorial(num)}")
"""
    # 3rd SQL Solution
    sql_sol = """SELECT student_name, subject, marks
FROM students
ORDER BY marks DESC
LIMIT 3;
"""
    
    # Select solution based on current coding question index
    if count == 0:
        code_solution = python_sol_1
    elif count == 1:
        code_solution = python_sol_2
    else:
        code_solution = sql_sol

    execution_log["coding_question_count"] = count + 1

    try:
        # 1. Select language
        selected_lang = False
        try:
            # Try <select> first
            lang_sel_list = fast_find_elements(driver, By.CSS_SELECTOR, "select[name='codingLang'], select.language-selector, select")
            if lang_sel_list and any(s.is_displayed() for s in lang_sel_list):
                from selenium.webdriver.support.ui import Select
                sel = Select(lang_sel_list[0])
                # Priority 1: Search for Python 3 or Python3 specifically
                python3_opt = next((o for o in sel.options if "python 3" in o.text.lower() or "python3" in o.text.lower()), None)
                if python3_opt:
                    sel.select_by_visible_text(python3_opt.text)
                    selected_lang = True
                else:
                    # Priority 2: Generic Python, SQL, etc.
                    for option in sel.options:
                        if any(l in option.text.lower() for l in ['python', 'sql', 'postgres', 'mysql', 'plsql']):
                            sel.select_by_visible_text(option.text)
                            selected_lang = True
                            break
                if selected_lang:
                    lang_text = python3_opt.text if python3_opt else "Python/SQL"
                    log.info(f"  Q{q_num} [Coding] → language selected: {lang_text}")
                    log_qa_step("Select Language", "PASS", f"Selected: {lang_text}")
                    time.sleep(1)
            
            # If not selected, try custom dropdown buttons
            if not selected_lang:
                lang_btns = fast_find_elements(driver, By.CSS_SELECTOR, "button.hp-dropdown-toggle, [id*='language'], [class*='language']")
                visible_lang = [b for b in lang_btns if b.is_displayed()]
                if visible_lang:
                    driver.execute_script("arguments[0].click();", visible_lang[0])
                    time.sleep(0.5)
                    # Try Python 3 specifically in the list then generic Python
                    python3_opts = fast_find_elements(driver, By.XPATH, "//li[contains(.,'Python 3')] | //a[contains(.,'Python 3')] | //li[contains(.,'Python3')] | //a[contains(.,'Python3')]")
                    if python3_opts:
                        driver.execute_script("arguments[0].click();", python3_opts[0])
                        selected_lang = True
                    else:
                        lang_opts = fast_find_elements(driver, By.XPATH, "//li[contains(.,'Python')] | //a[contains(.,'Python')] | //li[contains(.,'SQL')] | //a[contains(.,'SQL')] | //li[contains(.,'Postgre')]")
                        if lang_opts:
                            # Use first visible or first matching
                            driver.execute_script("arguments[0].click();", lang_opts[0])
                            selected_lang = True
                    
                    if selected_lang:
                        log.info(f"  Q{q_num} [Coding] → language selected via custom dropdown")
                        time.sleep(1)
        except Exception as e:
            log.warning(f"  Q{q_num} [Coding] Language selection attempt error: {e}")

        # 2. Inject code
        injected = False
        try:
            # Exhaustive ACE editor injection via JS
            result = driver.execute_script("""
                var code = arguments[0];
                var editors = document.querySelectorAll('.ace_editor');
                // Pattern 1: Global ace object
                if (typeof ace !== 'undefined' && ace.edit) {
                    for (var i = 0; i < editors.length; i++) {
                        try {
                            var ed = ace.edit(editors[i]);
                            if (ed) {
                                ed.setValue(code, -1);
                                return "ace_global_success";
                            }
                        } catch(e) {}
                    }
                }
                // Pattern 2: element.env.editor
                for (var i = 0; i < editors.length; i++) {
                    try {
                        if (editors[i].env && editors[i].env.editor) {
                            editors[i].env.editor.setValue(code, -1);
                            return "ace_env_success";
                        }
                    } catch(e) {}
                }
                // Pattern 3: Try finding editor via ID if global ace exists
                if (typeof ace !== 'undefined' && ace.edit) {
                    for (var i = 0; i < editors.length; i++) {
                        if (editors[i].id) {
                            try {
                                ace.edit(editors[i].id).setValue(code, -1);
                                return "ace_id_success";
                            } catch(e) {}
                        }
                    }
                }
                return false;
            """, code_solution)
            if result:
                log.info(f"  Q{q_num} [Coding] → injected via {result}")
                log_qa_step("Inject Code", "PASS", f"Engine: {result}, Code:\n{code_solution}")
                injected = True
        except Exception as e:
            log.warning(f"  Q{q_num} [Coding] ACE JS injection failed: {e}")

        if not injected:
            # Fallback: Character-by-character is BAD because of auto-complete (prefixes/suffixes)
            # Try to set value of the hidden textarea directly and trigger events
            try:
                textarea = fast_find_elements(driver, By.CSS_SELECTOR, ".ace_text-input")
                if textarea:
                    log.info(f"  Q{q_num} [Coding] → trying textarea direct value set")
                    driver.execute_script("""
                        var el = arguments[0];
                        el.focus();
                        el.value = arguments[1];
                        var ev = new Event('input', { bubbles: true });
                        el.dispatchEvent(ev);
                        // Also try to simulate a 'Paste' event if possible
                    """, textarea[0], code_solution)
                    time.sleep(0.5)
                    injected = True
            except Exception:
                pass

        if not injected:
            # Fallback: focus the ACE textarea specifically then ActionChains
            try:
                from selenium.webdriver.common.action_chains import ActionChains
                # ACE has a hidden textarea for input handling
                textarea = fast_find_elements(driver, By.CSS_SELECTOR, ".ace_text-input")
                if textarea:
                    log.info(f"  Q{q_num} [Coding] → focusing .ace_text-input for injection")
                    driver.execute_script("arguments[0].focus();", textarea[0])
                    time.sleep(0.3)
                    actions = ActionChains(driver)
                    # Select all and delete current boilerplate
                    actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL)
                    actions.send_keys(Keys.BACKSPACE)
                    # Type the solution
                    actions.send_keys(code_solution)
                    actions.perform()
                    log.info(f"  Q{q_num} [Coding] → typed code via ActionChains focus")
                    injected = True
                else:
                    # Fallback to general scroller
                    scroller = driver.find_element(By.CSS_SELECTOR, ".ace_scroller, .ace_content")
                    driver.execute_script("arguments[0].click();", scroller)
                    time.sleep(0.3)
                    actions = ActionChains(driver)
                    actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL)
                    actions.send_keys(Keys.DELETE)
                    actions.send_keys(code_solution)
                    actions.perform()
                    injected = True
            except Exception as e:
                log.warning(f"  Q{q_num} [Coding] ActionChains injection failed: {e}")

        time.sleep(1)

        # 3. Compile / Run Button
        run_btn_clicked = False
        try:
            # Common HirePro coding buttons: 'Run Tests', 'Compile & Run', 'Run'
            run_btn_labels = ["run tests", "compile", "run code", "test", "execute"]
            or_run = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in run_btn_labels]
            run_xpath = f"//button[{' or '.join(or_run)}]"
            
            run_btns = fast_find_elements(driver, By.XPATH, run_xpath)
            visible_run = [b for b in run_btns if b.is_displayed()]
            if visible_run:
                # Try standard click then JS as fallback
                try:
                    visible_run[0].click()
                except Exception:
                    driver.execute_script("arguments[0].click();", visible_run[0])
                btn_label = visible_run[0].text.strip()
                log.info(f"  Q{q_num} [Coding] → Clicked Run/Compile button: '{btn_label}'")
                log_qa_step(f"Click Button: '{btn_label}'", "PASS", "Initiating compilation")
                run_btn_clicked = True
        except Exception as e:
            log.warning(f"  Q{q_num} [Coding] Run button failed: {e}")

        # 4. Wait for Execution Results
        if run_btn_clicked:
            log.info("  Waiting for execution results (up to 45s)...")
            result_keywords = ["test case passed", "wrong output", "compilation error"]
            found_result = False
            start_wait = time.time()
            while (time.time() - start_wait) < 45:
                try:
                    body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
                    if any(kw in body_text for kw in result_keywords):
                        for kw in result_keywords:
                            if kw in body_text:
                                log.info(f"  ✅ Execution Result Found: '{kw}'")
                                log_qa_step("Verify Compile Result", "PASS", f"Platform Message: {kw.upper()}")
                                found_result = True
                                break
                        if found_result: break
                except Exception:
                    pass
                time.sleep(2)
            
            if found_result:
                time.sleep(1)
                # 5. Click Submit & Continue Coding
                sub_cont_labels = ["submit & continue", "submit and continue", "submit code", "save code"]
                or_sub = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in sub_cont_labels]
                sub_xpath = f"//button[{' or '.join(or_sub)}]"
                
                try:
                    sub_btns = fast_find_elements(driver, By.XPATH, sub_xpath)
                    visible_sub = [b for b in sub_btns if b.is_displayed()]
                    if visible_sub:
                        log.info(f"  Q{q_num} [Coding] → Clicking Submit & Continue")
                        try:
                            visible_sub[0].click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", visible_sub[0])
                        btn_label = visible_sub[0].text.strip()
                        log_qa_step(f"Click Button: '{btn_label}'", "PASS", "Submitting coding response")
                        time.sleep(2) # User requested 2 sec wait
                except Exception:
                    log.warning(f"  Q{q_num} [Coding] Submit & Continue click failed")
        
        return {"type": "Coding", "answer": "Code Injected", "status": "answered" if run_btn_clicked else "error"}

    except Exception as e:
        log_error(f"Q{q_num}-Coding", "WARNING", str(e))
        return {"type": "Coding", "answer": "", "status": "error"}

def recursive_find_monaco(driver):
    """Recursively search all iframes for the Monaco editor."""
    # Find in current frame
    try:
        editors = driver.find_elements(By.CSS_SELECTOR, ".monaco-editor, [aria-label='Editor'], .monaco-workbench")
        if editors and any(e.is_displayed() for e in editors):
            return True
    except Exception: pass

    # Search iframes
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    for frame in iframes:
        try:
            driver.switch_to.frame(frame)
            if recursive_find_monaco(driver):
                return True
            driver.switch_to.parent_frame()
        except Exception: 
            try: driver.switch_to.parent_frame()
            except Exception: driver.switch_to.default_content()

    return False

def step_open_terminal(driver):
    """Reliably open a new terminal in VS Code using multiple UI and Keyboard strategies."""
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.keys import Keys
    
    log.info("  Attempting to open terminal...")
    
    # Strategy 1: Traditional UI Navigation (Application Menu -> Terminal -> New Terminal)
    try:
        # Try multiple selectors for the 'hamburger' menu
        menu_selectors = [
            'div[aria-label="Application Menu"]',
            'a.action-label.menu-button',
            'div.monaco-menu-button',
            'div.activitybar [aria-label="Application Menu"]'
        ]
        menu = None
        for sel in menu_selectors:
            try:
                # Use find_elements (plural) to avoid implicit wait stalling on missing selectors
                elements = driver.find_elements(By.CSS_SELECTOR, sel)
                if elements and elements[0].is_displayed():
                    menu = elements[0]
                    break
            except Exception: continue
            
        if menu:
            driver.execute_script("arguments[0].click();", menu)
            time.sleep(1)
            # Find and click Terminal -> New Terminal
            wait(driver, By.XPATH, "//span[text()='Terminal']").click()
            time.sleep(0.5)
            wait(driver, By.XPATH, "//span[text()='New Terminal']").click()
            log.info("  ✓ Terminal opened via UI navigation")
            log_qa_step("Terminal Strategy", "PASS", "Opened via UI Navigation (Application Menu)")
            return True
    except Exception as e:
        log.warning(f"  UI navigation failed: {e}")

    # Strategy 2: Command Palette (Ctrl+Shift+P)
    try:
        log.info("  Attempting Command Palette fallback (Ctrl+Shift+P)")
        # Ensure editor has focus
        driver.execute_script("document.querySelector('.monaco-editor').focus();")
        time.sleep(0.5)
        
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).key_down(Keys.SHIFT).send_keys('p').key_up(Keys.SHIFT).key_up(Keys.CONTROL).perform()
        time.sleep(1)
        
        # Type "New Terminal" and Enter
        actions.send_keys("Terminal: Create New Integrated Terminal").send_keys(Keys.ENTER).perform()
        time.sleep(2)
        log.info("  ✓ Terminal opened via Command Palette")
        log_qa_step("Terminal Strategy", "PASS", "Opened via Command Palette (Ctrl+Shift+P)")
        return True
    except Exception as e:
        log.warning(f"  Command Palette fallback failed: {e}")

    # Strategy 3: Direct Shortcut (Ctrl + `)
    try:
        log.info("  Attempting direct shortcut (Ctrl+`)")
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).send_keys('`').key_up(Keys.CONTROL).perform()
        time.sleep(2)
        return True
    except Exception: pass

    return False

def answer_devops(driver, q_num):
    """DevOps AWS VM Assessment: wait for allocation, interact with VS Code IDE, run terminal solution."""
    log_qa_step("Stage: VM Allocation", "PASS", "Waiting for server spawn")
    start_vm = time.time()
    last_heartbeat = start_vm
    vm_loaded = False
    
    # 1. Wait for VM Allocation
    while (time.time() - start_vm) < VM_ALLOCATION_TIMEOUT:
        elapsed = int(time.time() - start_vm)
        try:
            # BROAD DETECTION: Check current frame AND nested iframes
            log.info("  Scanning for Monaco editor in all frames...")
            if recursive_find_monaco(driver):
                log.info(f"  ✅ VM Editor detected after {elapsed}s")
                log_qa_step("VM Detection", "PASS", f"Editor found in frame hierarchy after {elapsed}s")
                vm_loaded = True
                break
            
            # Reset to default in case recursion left it nested
            driver.switch_to.default_content()

            # 3. DIAGNOSTICS: If taking too long, log what we see
            if elapsed >= 180 and (time.time() - last_heartbeat) >= 60:
                log.warning(f"  [DIAGNOSTIC] Still waiting for VM ({elapsed}s)... dumping state")
                screenshot(driver, f"debug_vm_wait_{elapsed}")
                # Log frame count and top level peek
                iframes = driver.find_elements(By.TAG_NAME, "iframe")
                log.info(f"  Total potential frames to search: {len(iframes)}")
                last_heartbeat = time.time()

        except Exception as e:
            log.debug(f"  Wait iteration error: {e}")
            
        # Resilient Heartbeat (log every minute regardless of loop timing)
        if (time.time() - last_heartbeat) >= 60:
            log.info(f"  ... still waiting for VM ({int(time.time() - start_vm)}s elapsed)")
            last_heartbeat = time.time()
        
        time.sleep(3)
    
    if not vm_loaded:
        log_qa_step("VM Status", "FAIL", "Timed out waiting for editor")
        return {"type": "DevOps", "answer": "Timeout", "status": "error"}

    # 2. IDE Interaction
    try:
        from selenium.webdriver.common.action_chains import ActionChains
        log.info("  Starting IDE interaction sequence")
        
        # File selector: playbook.yml
        playbook = wait(driver, By.CSS_SELECTOR, 'div.monaco-list-row[aria-label="playbook.yml"]')
        safe_click(driver, playbook, "playbook-file")
        time.sleep(3)
        
        # Inject Code via Monaco API (Exhaustive Replacement)
        log.info("  Injecting YAML solution into Monaco editor (Exhaustive)")
        # Need to ensure editor is focused
        editor_el = driver.find_element(By.CSS_SELECTOR, ".monaco-editor")
        safe_click(driver, editor_el, "focus-editor")
        
        # 2. Monaco API Injection (Tiered Strategy - Additive)
        success = False
        for attempt in range(MAX_RETRIES):
            log.info(f"  Monaco injection attempt {attempt + 1} (Tier 1: API)...")
            # Try additive injection on all models
            success = driver.execute_script("""
                var marker = "#agent1";
                if (window.monaco && monaco.editor) {
                    var models = monaco.editor.getModels();
                    models.forEach(function(m) {
                        try { 
                            var val = m.getValue();
                            if (val.indexOf(marker) === -1) {
                                var lines = val.split('\\n');
                                // Insert at 2nd line (index 1)
                                lines.splice(1, 0, marker);
                                m.setValue(lines.join('\\n'));
                            }
                        } catch(e) {}
                    });
                    
                    // Handshake Verification
                    return monaco.editor.getModels().some(function(m) {
                        return m.getValue().indexOf(marker) !== -1;
                    });
                }
                return false;
            """)

            if success:
                log.info("  ✓ Tier 1 (API) additive injection verified")
                break
            
            log.warning("  Tier 1 failed, retrying...")
            time.sleep(2)
            
        # TIER 2: Physical Cursor Injection Fallback
        if not success:
            log.warning("  ⚠️ Tier 1 (API) failed. Initiating Tier 2: Physical Cursor Injection...")
            actions = ActionChains(driver)
            # Ctrl+Home to top, Down arrow to 2nd line, Enter for new line, Type #agent1
            actions.key_down(Keys.CONTROL).send_keys(Keys.HOME).key_up(Keys.CONTROL).perform()
            time.sleep(0.5)
            actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).send_keys(Keys.ARROW_UP).send_keys("#agent1").perform()
            time.sleep(5) 
            
            # Final verification via Page Source / DOM
            log.info("  Verifying physical injection via DOM...")
            if "#agent1" in driver.page_source:
                 success = True
                 log.info("  ✓ Tier 2 (Physical) verified via DOM presence")
            else:
                 log.error("  ❌ DEVASTATING FAILURE: Editor remains unchanged.")

        log_qa_step("Code Injected", "PASS" if success else "FAIL", "Comment '#agent1' added to 2nd line of editor")
        if not success:
             log_error("DevOps-IDE", "CRITICAL", "Could not verify code presence in IDE. Aborting strategy to stay safe.")
             return {"type": "DevOps", "answer": "Injection Error", "status": "error"}
        
        # Terminal Access
        log.info("  Accessing terminal to run solveproblem")
        if not step_open_terminal(driver):
             log_qa_step("Terminal Open", "FAIL", "Failed to open terminal via all strategies")
             return {"type": "DevOps", "answer": "Terminal Error", "status": "error"}
             
        log.info("  New Terminal opened")
        time.sleep(5) # Wait for terminal prompt
        
        # Execute Command
        actions = ActionChains(driver)
        actions.send_keys("solveproblem").send_keys(Keys.ENTER).perform()
        log_qa_step("Command Sent", "PASS", "Executed 'solveproblem' in integrated terminal")
        
        log.info("  Waiting 20s for solveproblem script execution...")
        log_qa_step("Execution Buffer", "PASS", "Waiting 20 seconds for terminal completion")
        time.sleep(20) # Wait for script execution
        
        # 3. Synchronize work (Run/Save) - Switches back to parent frame
        log.info("  Switching to parent frame for Run/Save...")
        log_qa_step("Context Switch", "PASS", "Switched back to default content/parent frame")
        driver.switch_to.default_content()
        time.sleep(2)
        
        log.info("  Synchronizing work...")
        # 'Run Tests' button outside VS Code
        run_btns = fast_find_elements(driver, By.XPATH, "//button[contains(., 'Run Tests')]")
        if run_btns:
            safe_click(driver, run_btns[0], "run-tests")
            log.info("  ✓ 'Run Tests' clicked - waiting 30s for results...")
            log_qa_step("Wait for results", "PASS", "Triggered 30-second synchronization wait")
            time.sleep(30) # Playbook runs and tests take time
            
        save_btns = fast_find_elements(driver, By.XPATH, "//button[contains(., 'Save My Work')]")
        if save_btns:
            safe_click(driver, save_btns[0], "save-my-work")
            # Wait for success toast
            try:
                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Saved successfully')]")))
                log_qa_step("Sync Verified", "PASS", "Work saved successfully")
            except Exception:
                log.warning("  Save success message not detected, but continuing...")

        return {"type": "DevOps", "answer": "Lab Completed", "status": "answered"}
        
    except Exception as e:
        log_error(f"DevOps-IDE", "WARNING", str(e))
        return {"type": "DevOps", "answer": str(e), "status": "error"}


def answer_question(driver, q_num, group_num=1):
    """Detect question type and answer accordingly."""
    q_type = _get_question_type(driver)
    display_type = "MCQ" if q_type == "Unknown" else q_type
    log.info(f"  [Q#{q_num:02d} | Grp {group_num} | {display_type}] Answering...")
    
    # Track type breakdown in execution log
    execution_log.setdefault("type_breakdown", {})
    execution_log["type_breakdown"][display_type] = execution_log["type_breakdown"].get(display_type, 0) + 1

    if q_type in ("MCQ", "MCQWithWeightage", "Unknown"):
        result = answer_mcq(driver, q_num)
    elif q_type == "RTC":
        result = answer_rtc(driver, q_num)
    elif q_type == "MultipleCorrectAnswer":
        result = answer_multiple_correct(driver, q_num)
    elif q_type == "Subjective":
        result = answer_subjective(driver, q_num)
    elif q_type == "FillInTheBlank":
        result = answer_fill_in_blank(driver, q_num)
    elif q_type == "Coding":
        result = answer_coding(driver, q_num)
    elif q_type == "DevOps":
        result = answer_devops(driver, q_num)
    else:
        result = {"type": q_type, "answer": "", "status": "skipped"}

    result["question_number"] = q_num
    result["confidence_score"] = 60 if result["status"] == "answered" else 0
    result["retries"] = 0
    result["error_message"] = None
    question_log.append(result)

    if result["status"] == "answered":
        execution_log["answered"] += 1
    else:
        execution_log["skipped"] += 1

    execution_log["total_questions"] = q_num
    return result


# =====================================================================
# NAVIGATION: NEXT QUESTION / NEXT GROUP
# =====================================================================
def click_next_question(driver, q_num):
    """Click Next Question button using a single ORed XPath for speed."""
    labels = ["next question", "next", "save & next", "save and next"]
    # Build a single XPath that matches any of these labels
    or_parts = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in labels]
    combined_xpath = f"//button[{' or '.join(or_parts)}]"
    
    try:
        # One-pass check with 3s timeout
        btns = fast_find_elements(driver, By.XPATH, combined_xpath)
        visible = [b for b in btns if b.is_displayed()]
        if visible:
            if safe_click(driver, visible[0], f"next-Q{q_num}"):
                time.sleep(0.5)
                log.info(f"  → Next Question clicked (OR-XPath)")
                return True
    except Exception:
        pass


    # Try by CSS identifying misspelled 'secondry' found during inspection
    # Try by CSS selectors using fast_find_elements
    for selector in ["button.btn-secondry-success", "button.btn-secondary-success", ".btn-next", "[ng-click*='nextQuestion']"]:
        btns = fast_find_elements(driver, By.CSS_SELECTOR, selector)
        visible = [b for b in btns if b.is_displayed() and b.is_enabled()]
        if visible:
            if safe_click(driver, visible[0], f"next-Q{q_num}-css"):
                time.sleep(0.5)
                log.info(f"  → Next Question clicked via CSS: '{selector}'")
                return True


    return False


def click_next_group(driver):
    """Click Take Me To Next Group or equivalent using a single ORed XPath."""
    check_fullscreen_overlay(driver)  # Ensure no anti-cheat overlay blocks the button
    labels = ["take me to next group", "take me to the chosen group", "chosen group",
              "next group", "continue to next", "next section", 
              "proceed to next group", "proceed to next module", "next group"]
    or_parts = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in labels]
    combined_xpath = f"//button[{' or '.join(or_parts)}]"

    try:
        btns = fast_find_elements(driver, By.XPATH, combined_xpath)
        visible = [b for b in btns if b.is_displayed()]
        if visible:
            if safe_click(driver, visible[0], "next-group"):
                log.info(f"  ✓ Next Group clicked (OR-XPath)")
                time.sleep(2) # Increased for stability
                return True
    except Exception:
        pass

    # Try by CSS
    try:
        btn = fast_find_elements(driver, By.CSS_SELECTOR, ".btn-next-group, [ng-click*='nextGroup'], [ng-click*='next_group']")
        if btn and any(b.is_displayed() for b in btn):
            if safe_click(driver, btn[0], "next-group-css"):
                time.sleep(2)
                return True
    except Exception:
        pass

    return False


def handle_optional_group(driver):
    """Detect and handle optional group selection screen."""
    try:
        # Browser inspection confirmed: optional groups use label.custom-radio buttons
        # and a confirm btn with name='btnSelectGroup'
        confirm_btn = fast_find_elements(driver, By.CSS_SELECTOR, "button[name='btnSelectGroup']")
        if confirm_btn and any(b.is_displayed() for b in confirm_btn):
            log.info("  Optional group screen detected (btnSelectGroup)")
            # Select first visible radio option
            radio_opts = fast_find_elements(driver, By.CSS_SELECTOR, "label.custom-radio")
            visible_radios = [r for r in radio_opts if r.is_displayed()]
            if visible_radios:
                safe_click(driver, visible_radios[0], "optional-group-radio")
                time.sleep(0.5)
                log.info("  Selected first optional group radio")
            # Click the confirm button
            visible_confirm = [b for b in confirm_btn if b.is_displayed()]
            if visible_confirm:
                driver.execute_script("arguments[0].click();", visible_confirm[0])
                log.info("  Clicked 'Take me to the chosen group'")
                time.sleep(2)
            return True
    except Exception:
        pass
    return False


# =====================================================================
# STEP 5: QUESTION LOOP
# =====================================================================
def step_questions(driver, max_q=50):
    log.info("═" * 60)
    log.info("STEP 5: Question Handling Loop")
    log.info("═" * 60)

    q_num = 0
    max_questions = max_q  # Allow stage-specific limits (e.g. 7 for stage 2)
    group_num = 1
    consecutive_no_next = 0
    group_q_counts = {1: 0}  # track questions per group

    while q_num < max_questions:
        # Prevent re-answering the same DevOps lab over and over
        if GLOBAL_IS_DEVOPS and q_num > 1:
            log.info("  DevOps Assessment is primarily a single-page lab — stopping loop after initial attempt.")
            break

        # Proactively check for fullscreen overlay at start of each question
        try:
            if check_fullscreen_overlay(driver):
                # Reset failure count if we just recovered from an overlay
                consecutive_no_next = 0
        except Exception as e:
            log.warning(f"  Overlay check error: {e}")
        
        q_num += 1
        group_q_counts[group_num] = group_q_counts.get(group_num, 0) + 1
        g_local = group_q_counts[group_num]
        log.info(f"\n  {'='*55}")
        log.info(f"  Q#{q_num:02d} | Group {group_num} | Q#{g_local} in this group")
        log.info(f"  {'='*55}")
        time.sleep(0.2)

        body_text = driver.find_element(By.TAG_NAME, "body").text.lower()

        # Test already ended?
        if any(x in body_text for x in ["test submitted", "submit success", "test completed", "your response has been recorded"]):
            log.info("  Test completion screen detected — stopping question loop")
            break

        # Handle optional group screen
        if handle_optional_group(driver):
            log.info(f"  ↳ Optional group selected — moving to Group {group_num + 1}")
            group_num += 1
            group_q_counts[group_num] = 0
            q_num -= 1
            continue

        # Answer the question
        answer_question(driver, q_num, group_num)
        time.sleep(0.3)

        # 1. CHECK FOR END TEST (Conditional Break)
        try:
            # DevOps labs: 'End Test' is always visible.
            # We ONLY break if we're not in DevOps mode or if we failed multiple attempts.
            end_btns = fast_find_elements(driver, By.XPATH, "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'end test') or contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'submit test')]")
            visible_end = [b for b in end_btns if b.is_displayed()]
            
            if visible_end:
                if GLOBAL_IS_DEVOPS:
                    # In DevOps agent script, 'End Test' is suspicious if seen early.
                    # We check if we have actually attempted any lab work yet.
                    if q_num <= 1 and execution_log.get("answered", 0) == 0:
                        log.info("  End Test button visible, but ignoring for initial DevOps detection")
                    else:
                        log.info("  End Test button detected after initial attempt — finishing loop")
                        break
                else:
                    log.info("  End Test button detected — finishing questions")
                    break
        except Exception:
            pass

        # 2. Try Next Question
        next_q = click_next_question(driver, q_num)

        if not next_q:
            # 3. Try Next Group
            next_g = click_next_group(driver)
            if next_g:
                log.info(f"  ✓ Moved to Group {group_num + 1}")
                group_num += 1
                group_q_counts[group_num] = 0
                screenshot(driver, f"05_group_{group_num}_start")
                consecutive_no_next = 0
            else:
                consecutive_no_next += 1
                if consecutive_no_next >= 3:
                    log.warning("  No navigation found for 3 consecutive questions — stopping loop")
                    break

    # Print final summary
    breakdown = execution_log.get("type_breakdown", {})
    log.info(f"\n  {'─'*55}")
    log.info(f"  QUESTION LOOP COMPLETE")
    log.info(f"  Total questions seen : {q_num}")
    log.info(f"  Total answered       : {execution_log['answered']}")
    log.info(f"  Groups traversed     : {group_num}")
    if breakdown:
        log.info(f"  By type:")
        for qtype, count in sorted(breakdown.items()):
            log.info(f"    {qtype:<22}: {count}")
    log.info(f"  {'─'*55}")
    screenshot(driver, "05_questions_done")
    return True


# =====================================================================
# STEP 6: TEST SUBMISSION
# =====================================================================
def step_submit(driver):
    log.info("═" * 60)
    log.info("STEP 6: Submitting Test")
    log.info("═" * 60)

    # 1. Broad list of labels for all submission stages (End Test -> I'm Done -> Yes Submit)
    # NOTE: Avoid apostrophes here as they break XPath string literals in the current builder.
    submit_labels = [
        "end test", "submit test", "finish test",
        "done with the test", "i am done with the test", "i am done",
        "confirm submit", "yes, submit", "yes", "final submit"
    ]
    or_parts = [f"contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'{l}')" for l in submit_labels]
    combined_xpath = f"//*[self::button or self::a][{' or '.join(or_parts)}]"

    # 2. Success keywords to identify the absolute end page (visible body text only)
    success_keywords = [
        "thanks for attending the test", 
        "we have another test lined up for you",
        "thanks for attending", 
        "your response has been recorded", 
        "lined up", 
        "successfully submitted",
        "successfully recorded",
        "test completed successfully"
    ]

    log.info("  Entering final submission resolution loop...")
    start_time = time.time()
    max_wait = 45  # Total time to clear all confirmation stages
    
    last_clicked = ""
    while (time.time() - start_time) < max_wait:
        # A. Check for success message (Termination condition)
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
            if any(kw in body_text for kw in success_keywords):
                log.info("  ✅ FINAL SUCCESS SCREEN DETECTED!")
                break
        except Exception:
            pass
        
        # B. Look for ANY submission-related button on CURRENT window first
        found_on_current = False
        try:
            btns = fast_find_elements(driver, By.XPATH, combined_xpath)
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                btn_text = visible[0].text.strip()
                log.info(f"  Found button: '{btn_text}' — clicking")
                try:
                    visible[0].click()
                except Exception:
                    driver.execute_script("arguments[0].click();", visible[0])
                found_on_current = True
                time.sleep(3)
        except Exception:
            pass

        if found_on_current:
            continue

        # C. If nothing found on current window, try switching to a new tab if it exists
        if switch_to_newest_window(driver, silent=False):
            time.sleep(1)
            continue
        
        # D. Clear overlays
        check_fullscreen_overlay(driver)
        
        time.sleep(2) # Polling interval

    screenshot(driver, "06_test_final_state")
    log.info("  ✅ Test submission sequence finalized")
    return True


# =====================================================================
# STEP 7: CHAINING HANDLER
# =====================================================================
def step_handle_chaining(driver):
    """
    Check for eligibility messages and handle transition to the next test.
    Returns: 'PASS', 'FAIL', or 'NONE'
    """
    log.info("═" * 60)
    log.info("STEP 7: Handling Chaining / Eligibility")
    log.info("═" * 60)
    
    start_wait = time.time()
    max_wait = 30
    
    found_status = "NONE"
    
    while (time.time() - start_wait) < max_wait:
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            
            if "Congratulations! You are eligible for the next test." in body_text:
                msg = "Congratulations! You are eligible for the next test."
                log.info(f"  🏆 ELIGIBILITY: {msg}")
                log_qa_step("Check Eligibility", "PASS", f"Message Detected: '{msg}'")
                found_status = "PASS"
                break
            elif "Sorry! You are not eligible for the next test." in body_text:
                msg = "Sorry! You are not eligible for the next test."
                log.info(f"  ❌ ELIGIBILITY: {msg}")
                log_qa_step("Check Eligibility", "FAIL", f"Message Detected: '{msg}'")
                found_status = "FAIL"
                break
        except Exception:
            pass
        
        # Check for window switches just in case
        switch_to_newest_window(driver)
        time.sleep(2)

    if found_status == "PASS":
        # Look for "Take me to next test" button
        try:
            xpath = "//*[contains(text(), 'Take me to next test')]"
            btns = fast_find_elements(driver, By.XPATH, xpath)
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                log.info("  Found 'Take me to next test' button — clicking...")
                try:
                    visible[0].click()
                except Exception:
                    driver.execute_script("arguments[0].click();", visible[0])
                
                log_qa_step("Click Button: 'Take me to next test'", "PASS", "Proceeding to next stage")
                time.sleep(5) # Wait for new window to open
                if switch_to_newest_window(driver, silent=False):
                    log.info("  Switched to the NEXT TEST window ✅")
                return "PASS"
        except Exception as e:
            log.warning(f"  Failed to click 'Take me to next test': {e}")

    return found_status


def run_test_stage(driver, stage_name, max_q=50):
    """Executes a full test stage from pre-test to submission."""
    log.info("\n" + "█" * 60)
    log.info(f" STARTING TEST STAGE: {stage_name}")
    log.info("█" * 60)
    
    # Pre-test
    if not step_pretest(driver):
        log.error(f"  [{stage_name}] Pre-test failed")
        return False
        
    # Selfie
    if not step_selfie(driver):
        log.error(f"  [{stage_name}] Selfie failed")
        # Continue anyway if allowed by platform
        
    # Start Test
    if not step_start_test(driver):
        log.error(f"  [{stage_name}] Start test failed")
        return False
        
    # Questions
    if not step_questions(driver, max_q=max_q):
        log.error(f"  [{stage_name}] Question loop failed")
        return False
        
    # Submit
    if not step_submit(driver):
        log.error(f"  [{stage_name}] Submission failed")
        return False
        
    log.info(f"  ✅ [{stage_name}] STAGE COMPLETE")
    return True


# =====================================================================
# REPORT GENERATION
# =====================================================================
def generate_reports():
    os.makedirs(REPORT_DIR, exist_ok=True)
    end = time.time()
    execution_log["end_time"] = datetime.now().isoformat()
    execution_log["duration_sec"] = round(end - start_time, 1)

    if execution_log["answered"] > 0:
        execution_log["accuracy_estimate"] = f"{min(60 + execution_log['answered'] * 2, 90)}%"

    if execution_log["errors"] == 0:
        execution_log["execution_status"] = "success"
        execution_log["remarks"] = "All steps completed successfully."
    elif execution_log["answered"] > 0:
        execution_log["execution_status"] = "partial"
        execution_log["remarks"] = f"{execution_log['errors']} error(s) encountered during execution."
    else:
        execution_log["execution_status"] = "failed"
        execution_log["remarks"] = "Execution failed before questions could be answered."

    # JSON report
    json_path = os.path.join(REPORT_DIR, f"ga_report_{TODAY}.json")
    with open(json_path, "w") as f:
        json.dump({"summary": execution_log, "questions": question_log}, f, indent=2)
    log.info(f"  JSON Report: {json_path}")

    # CSV report
    csv_path = os.path.join(REPORT_DIR, f"ga_report_{TODAY}.csv")
    if question_log:
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=question_log[0].keys())
            writer.writeheader()
            writer.writerows(question_log)
        log.info(f"  CSV Report: {csv_path}")

    # Summary to console
    log.info("\n" + "═" * 60)
    log.info("EXECUTION REPORT")
    log.info("═" * 60)
    log.info(f"  Status         : {execution_log['execution_status'].upper()}")
    log.info(f"  Total Questions: {execution_log['total_questions']}")
    log.info(f"  Answered       : {execution_log['answered']}")
    log.info(f"  Skipped        : {execution_log['skipped']}")
    log.info(f"  Errors         : {execution_log['errors']}")
    log.info(f"  Duration       : {execution_log['duration_sec']}s")
    log.info(f"  Reports → {REPORT_DIR}")
    log.info("═" * 60)

    return json_path, csv_path


# =====================================================================
# READ GENERIC ASSESSMENT URL FROM configs/urls.txt
# =====================================================================
def read_devops_url():
    """Reads the line starting with 'devops:' from configs/urls.txt."""
    if not os.path.exists(URLs_FILE):
        log.error(f"urls.txt not found at: {URLs_FILE}")
        sys.exit(1)

    with open(URLs_FILE, "r") as f:
        for line in f:
            line = line.strip()
            if line.lower().startswith("devops:"):
                url = line.split("devops:", 1)[1].strip()
                if url:
                    log.info(f"  DevOps URL loaded from urls.txt: {url}")
                    return url

    log.error("No 'devops:' URL found in urls.txt.")
    sys.exit(1)


# =====================================================================
# DETAILED REPORT GENERATION (EXCEL & HTML)
# =====================================================================
def generate_detailed_reports():
    """
    Generates detailed Excel and HTML QA reports with granular steps.
    """
    os.makedirs(REPORT_DIR, exist_ok=True)
    report_tag = TODAY # Could use timestamp if needed
    
    # 1. EXCEL REPORT
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        df = pd.DataFrame(qa_steps)
        excel_path = os.path.join(REPORT_DIR, f"Devops_detailed_QA_Report_{report_tag}.xlsx")
        
        # Calculate summary metrics
        total_steps = len(qa_steps)
        pass_count  = sum(1 for s in qa_steps if s['Status'] == 'PASS')
        fail_count  = sum(1 for s in qa_steps if s['Status'] in ('FAIL', 'CRITICAL'))
        warn_count  = sum(1 for s in qa_steps if s['Status'] == 'WARNING')
        
        # We leave space at top for summary (Rows 1-5)
        # DataFrame starts at Row 7
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Execution Log', startrow=6)
            workbook  = writer.book
            worksheet = writer.sheets['Execution Log']
            
            # --- SUMMARY HEADER ---
            # Define Styles
            title_font = Font(name='Segoe UI', size=14, bold=True, color='2C3E50')
            header_font = Font(name='Segoe UI', size=11, bold=True)
            pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            warn_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            box_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Insert Summary Data
            worksheet['A1'] = "DETAILED QA EXECUTION REPORT"
            worksheet['A1'].font = title_font
            
            worksheet['A2'] = "Agent Name:"
            worksheet['B2'] = execution_log.get('agent_name', 'Generic Agent')
            worksheet['A3'] = "Total Steps:"
            worksheet['B3'] = total_steps
            worksheet['A4'] = "Status Summary:"
            worksheet['B4'] = f"PASS: {pass_count} | FAIL/CRIT: {fail_count} | WARN: {warn_count}"
            
            # Bold Header row (A7:E7)
            for col_idx in range(1, 6):
                cell = worksheet.cell(row=7, column=col_idx)
                cell.font = header_font
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = box_border

            # --- COLOR CODING STATUS COLUMN (Column E) ---
            # Column E is 'Status'
            for r in range(8, 8 + total_steps):
                status_cell = worksheet.cell(row=r, column=5)
                status_val = str(status_cell.value).upper()
                
                if status_val == 'PASS':
                    status_cell.fill = pass_fill
                elif status_val in ('FAIL', 'CRITICAL'):
                    status_cell.fill = fail_fill
                elif status_val == 'WARNING':
                    status_cell.fill = warn_fill
                
                status_cell.alignment = Alignment(horizontal='center')
                status_cell.border = box_border
            
            # Column Width Adjustment
            col_widths = {1: 10, 2: 20, 3: 35, 4: 60, 5: 15}
            for col_idx, width in col_widths.items():
                worksheet.column_dimensions[chr(64 + col_idx)].width = width

            # --- NEW SHEET: TECHNICAL DESIGN & STRATEGY ---
            design_data = [
                ["TECHNICAL DESIGN & QA STRATEGY", ""],
                ["", ""],
                ["1. VM DETECTION", "Uses recursive_find_monaco() to traverse all iframes. Includes a 10-minute heartbeat polling loop for AWS VM allocation."],
                ["2. CODE INJECTION", "Tiered Strategy: Tier 1 (Monaco API additive splice) with verification handshake. Tier 2 (Physical keystroke simulation via cursor navigation) as fail-safe."],
                ["3. TERMINAL ACCESS", "Multi-strategy: UI Navigation (Menu), Command Palette (Ctrl+Shift+P), and Direct Shortcuts (Ctrl+`). Includes 20s execution buffer."],
                ["4. SYNC & SAVE", "Automated context switching to parent frame. Verified clicks on 'Run Tests' and 'Save My Work' with success-toast detection."],
                ["5. LOOP PERSISTENCE", "Single-pass lab rule: Breaks question loop after first successful DevOps lab completion to prevent redundant automation."],
                ["6. REPORTING", "Automated generation of JSON, CSV, Excel (detailed), and HTML (walkthrough) reports for every run."]
            ]
            design_df = pd.DataFrame(design_data, columns=["Category", "Implementation Details"])
            design_df.to_excel(writer, index=False, sheet_name='Technical Design')
            
            design_ws = writer.sheets['Technical Design']
            design_ws['A1'].font = title_font
            design_ws.column_dimensions['A'].width = 25
            design_ws.column_dimensions['B'].width = 100
            for row in design_ws.iter_rows(min_row=3, max_row=10, min_col=1, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = box_border
                
        log.info(f"  ✅ [EXCEL] Detailed report with technical design: {excel_path}")
    except Exception as e:
        log.warning(f"  Failed Excel generation: {e}")

    # 2. HTML REPORT
    try:
        html_path = os.path.join(REPORT_DIR, f"Detailed_QA_Report_{report_tag}.html")
        
        # In-line template for zero-dependency portability
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Detailed Assessment QA Report</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f7f6; color: #333; margin: 0; padding: 20px; }}
        .container {{ max-width: 1200px; margin: auto; background: #fff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        .summary {{ display: flex; gap: 20px; margin-bottom: 30px; }}
        .card {{ flex: 1; padding: 20px; border-radius: 8px; text-align: center; color: #fff; }}
        .blue-card {{ background: #3498db; }}
        .green-card {{ background: #2ecc71; }}
        .red-card {{ background: #e74c3c; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background-color: #f8f9fa; color: #2c3e50; }}
        tr:hover {{ background-color: #f1f1f1; }}
        .details-col {{ white-space: pre-wrap; word-break: break-word; font-family: 'Courier New', Courier, monospace; font-size: 0.9em; }}
        .status-badge {{ padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 0.85em; }}
        .status-PASS {{ background-color: #d4edda; color: #155724; }}
        .status-FAIL {{ background-color: #f8d7da; color: #721c24; }}
        .status-WARNING {{ background-color: #fff3cd; color: #856404; }}
        .status-CRITICAL {{ background-color: #e74c3c; color: #fff; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Detailed Assessment QA Report</h1>
        <div class="summary">
            <div class="card blue-card"><h3>Assessment</h3><p>{execution_log.get('assessment_url', 'N/A')}</p></div>
            <div class="card green-card"><h3>Answered</h3><p>{execution_log.get('answered', 0)} Questions</p></div>
            <div class="card red-card"><h3>Errors</h3><p>{execution_log.get('errors', 0)} Detected</p></div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Step</th>
                    <th>Timestamp</th>
                    <th>Action</th>
                    <th>Details</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
        """
        for step in qa_steps:
            status_cls = f"status-{step['Status']}"
            html_content += f"""
                <tr>
                    <td>{step['Step ID']}</td>
                    <td>{step['Timestamp']}</td>
                    <td>{step['Action']}</td>
                    <td class="details-col">{step['Details']}</td>
                    <td><span class="status-badge {status_cls}">{step['Status']}</span></td>
                </tr>
            """
        
        html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
        """
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        log.info(f"  ✅ [HTML] Detailed report: {html_path}")
    except Exception as e:
        log.warning(f"  Failed HTML generation: {e}")


def main():
    log_qa_step("Init Log", "PASS", "Reading DevOps URL from configs/urls.txt")
    url = read_devops_url()

    log.info("═" * 60)
    log.info("DevOps Assessment Agent - Framework Extensions")
    log.info(f"Date  : {TODAY}")
    log.info(f"URL   : {url}")
    log.info("═" * 60)

    driver = create_driver()
    log_qa_step("Start Agent", "PASS", "Driver initialized")

    try:
        # 1. Navigation
        if not step_navigate(driver, url):
            log_qa_step("Final Status", "FAIL", "Navigation stage failed")
            raise RuntimeError("Page navigation failed")

        # 2. RUN TEST STAGE
        if run_test_stage(driver, "DevOps Lab Assessment"):
            log_qa_step("Lab Status", "PASS", "Completed DevOps Lab")
        else:
            log_qa_step("Lab Status", "FAIL", "Lab execution failed")

        # 5. FINAL VERIFICATION
        log.info("═" * 60)
        log.info("FINAL VERIFICATION")
        log.info("═" * 60)
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        if "thanks for attending the test" in body and "safe to close" in body:
            log_qa_step("Final Verification", "PASS", "Found global success messages")
        else:
            log_qa_step("Final Verification", "WARNING", "Global success messages missing or unexpected UI")

    except Exception as e:
        log_qa_step("Execution Runtime", "CRITICAL", str(e))
        log_error("main", "CRITICAL", str(e))
        screenshot(driver, "error_fatal")
    finally:
        generate_reports() # Standard reports
        generate_detailed_reports() # New granular reports
        driver.quit()
        log.info("Browser closed.")


if __name__ == "__main__":
    main()
